#!/usr/bin/env python
# -*- coding: utf-8 -*-

# export each worksheets as xlsx

from openpyxl import load_workbook
import os
from os import sys
import shutil

def get_all_files(in_dir):
    file_pathes = []
    for root, dirs, files in os.walk(in_dir):
        for file_name in files:
            file_pathes.append(''.join([in_dir, '/', file_name]))
    return file_pathes

def get_all_sheets(excel_file):
    sheets = []
    workbook = load_workbook(excel_file)
    all_worksheets = workbook.get_sheet_names()
    for worksheet_name in all_worksheets:
        sheets.append(worksheet_name)
    return sheets

def split_excel_sheet(excel_file, worksheet_names, out_dir):
    excel_file_name = excel_file.rsplit('/', 1)[1].split('.')[0];
    print("File " + excel_file_name + " ...")
    print("")
    for worksheet_name in worksheet_names:
        dir_path = ''.join([out_dir.decode('utf_8'), '/', excel_file_name.decode('utf_8')]);
        if not os.path.exists(dir_path):
            os.makedirs(dir_path);
        file_path = ''.join([dir_path, '/', worksheet_name,'.xlsx']);
        shutil.copyfile(excel_file, file_path)
        print("  Copy as: " + file_path + " ...")

        workbook = load_workbook(file_path,data_only=True)
        for worksheet in workbook.worksheets:
            if worksheet.title == worksheet_name:
                print("  Target Sheet: " + worksheet.title + " ...")
            else:
                workbook.remove(worksheet)
        workbook.save(file_path)

if len(sys.argv) != 3:
	print("Call with " + sys.argv[0] + " <in dir> <out dir>")
	sys.exit(1)
else:
    in_dir = sys.argv[1]
    out_dir = sys.argv[2]
    files = get_all_files(in_dir)
    print files
    for file_path in files:
        sheets = get_all_sheets(file_path)
        assert(sheets != None and len(sheets) > 0)
        split_excel_sheet(file_path, sheets, out_dir)
